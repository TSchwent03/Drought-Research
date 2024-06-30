from openpyxl import load_workbook
import os

def convert_text_to_numbers(filename):
    # Load the Excel file
    wb = load_workbook(filename)

    # Access the first sheet (modify sheet name if needed)
    sheet = wb.active

    # Iterate through all cells in the sheet
    for row in sheet.iter_rows():
        for cell in row:
            # Convert text to number if it's currently text
            if cell.value is not None and isinstance(cell.value, str):
                try:
                    cell.value = float(cell.value.replace(",", ""))  # Remove potential commas
                except ValueError:
                    pass  # Skip cells that cannot be converted to numbers

    # Save the modified file
    wb.save(filename)

# Define the directory containing your files
data_dir = r"C:\Users\thoma\Documents\GitHub\Drought-Research\Precip Data\Dataset B\Monthly Tot SPI w Drgt Stats Excel"

# Loop through all Excel files in the directory
for filename in os.listdir(data_dir):
    if filename.endswith(".xlsx") or filename.endswith(".xlsm"):  # Check for Excel file extensions
        convert_text_to_numbers(os.path.join(data_dir, filename))

print("Conversion completed successfully!")