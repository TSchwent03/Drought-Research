from openpyxl import load_workbook
import os

def convert_text_to_numbers(filename):
    # Load the Excel workbook
    workbook = load_workbook(filename)

    # Access default sheet (shoudln't be necessary but is here just in case)
    sheet = workbook.active

    # Iterate through all cells in the sheet
    for row in sheet.iter_rows():
        for cell in row:
            # Convert text to number if it's currently text
            if cell.value is not None and isinstance(cell.value, str):
                try:
                    cell.value = float(cell.value)
                except ValueError:
                    pass  # Skip cells that cannot be converted to numbers

    # Create new file
    new_filename = filename.replace(".xlsx", "_cleaned.xlsx")
    workbook.save(new_filename)

def main():
    # Define the directories
    program_path = os.path.dirname(os.path.abspath(__file__))
    rel_path = r"Precip Data\Dataset B-a\Monthly Tot SPI w Drgt Stats Excel"
    excel_path = os.path.abspath(os.path.join(program_path, "..", "..", rel_path))

    # Loop through all Excel files in the directory
    for filename in os.listdir(excel_path):
        if filename.endswith(".xlsx"):  # Check for Excel file extension
            convert_text_to_numbers(os.path.join(excel_path, filename))
    print("Conversion completed successfully!")

if __name__ == "__main__":
    main()